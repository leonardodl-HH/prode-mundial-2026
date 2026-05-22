import streamlit as st
import psycopg2
import pandas as pd
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import os
from datetime import datetime

# --- CONFIGURACIÓN DE PÁGINA STREAMLIT ---
st.set_page_config(page_title="Prode Mundial 2026", page_icon="⚽", layout="wide")

# --- CONEXIÓN DE ARQUITECTURA PROFESIONAL (POSTGRESQL CLOUD) ---
class DatabaseManager:
    @staticmethod
    def get_connection():
        return psycopg2.connect(st.secrets["postgres"]["url"])

    @staticmethod
    def init_db():
        conn = DatabaseManager.get_connection()
        
        comandos_tablas = [
            '''CREATE TABLE IF NOT EXISTS equipos (id_equipo SERIAL PRIMARY KEY, nombre TEXT UNIQUE NOT NULL, zona TEXT NOT NULL, archivo_bandera TEXT DEFAULT 'default.png')''',
            '''CREATE TABLE IF NOT EXISTS partidos (id_partido SERIAL PRIMARY KEY, fase TEXT NOT NULL, id_equipo_local INTEGER, id_equipo_visitante INTEGER, goles_local INTEGER DEFAULT NULL, goles_visitante INTEGER DEFAULT NULL, FOREIGN KEY(id_equipo_local) REFERENCES equipos(id_equipo), FOREIGN KEY(id_equipo_visitante) REFERENCES equipos(id_equipo), UNIQUE(fase, id_equipo_local, id_equipo_visitante))''',
            '''CREATE TABLE IF NOT EXISTS usuarios (id_usuario SERIAL PRIMARY KEY, nombre TEXT UNIQUE NOT NULL)''',
            '''CREATE TABLE IF NOT EXISTS apuestas (id_apuesta SERIAL PRIMARY KEY, id_usuario INTEGER, id_partido INTEGER, apuesta_goles_local INTEGER DEFAULT NULL, apuesta_goles_visitante INTEGER DEFAULT NULL, equipo_l_predicho TEXT, equipo_v_predicho TEXT, puntos_obtenidos REAL DEFAULT 0.0, FOREIGN KEY(id_usuario) REFERENCES usuarios(id_usuario), FOREIGN KEY(id_partido) REFERENCES partidos(id_partido), UNIQUE(id_usuario, id_partido))''',
            '''CREATE TABLE IF NOT EXISTS configuracion (id INTEGER PRIMARY KEY, pts_prode INTEGER, pts_exacto INTEGER, pts_parcial INTEGER, pts_dif INTEGER, api_key TEXT, id_liga TEXT, admin_pass TEXT, pts_prode_ko INTEGER DEFAULT 8, pts_exacto_ko INTEGER DEFAULT 6, pts_parcial_ko INTEGER DEFAULT 2, pts_dif_ko INTEGER DEFAULT 2, fecha_limite TEXT DEFAULT '2026-06-11 16:00:00')'''
        ]
        
        for cmd in comandos_tablas:
            try:
                cursor = conn.cursor()
                cursor.execute(cmd)
                conn.commit()
                cursor.close()
            except Exception:
                conn.rollback()
        
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO configuracion (id, pts_prode, pts_exacto, pts_parcial, pts_dif, api_key, id_liga, admin_pass, pts_prode_ko, pts_exacto_ko, pts_parcial_ko, pts_dif_ko, fecha_limite) 
                VALUES (1, 4, 3, 1, 1, '', '1', 'admin123', 8, 6, 2, 2, '2026-06-11 16:00:00')
                ON CONFLICT (id) DO NOTHING
            """)
            conn.commit()
            cursor.close()
        except Exception:
            conn.rollback()
        
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM equipos")
            if cursor.fetchone()[0] == 0:
                mapa_grupos = {
                    "A": ["México", "Estados Unidos", "Canadá", "Costa Rica"], "B": ["Argentina", "Brasil", "Uruguay", "Colombia"],
                    "C": ["Francia", "Inglaterra", "España", "Alemania"], "D": ["Portugal", "Italia", "Países Bajos", "Bélgica"],
                    "E": ["Croacia", "Marruecos", "Japón", "Senegal"], "F": ["Ecuador", "Perú", "Chile", "Paraguay"],
                    "G": ["Ghana", "Camerún", "Túnez", "Argelia"], "H": ["Corea del Sur", "Australia", "Irán", "Arabia Saudita"],
                    "I": ["Jamaica", "Panamá", "Honduras", "El Salvador"], "J": ["Nigeria", "Costa de Marfil", "Egipto", "Mali"],
                    "K": ["Suiza", "Dinamarca", "Serbia", "Ucrania"], "L": ["Polonia", "Suecia", "Escocia", "Gales"]
                }
                iso_mapping = {
                    "México": "mx", "Estados Unidos": "us", "Canadá": "ca", "Costa Rica": "cr", "Argentina": "ar", "Brasil": "br", "Uruguay": "uy", "Colombia": "co",
                    "Francia": "fr", "Inglaterra": "gb-eng", "España": "es", "Alemania": "de", "Portugal": "pt", "Italia": "it", "Países Bajos": "nl", "Bélgica": "be",
                    "Croacia": "hr", "Marruecos": "ma", "Japón": "jp", "Senegal": "sn", "Ecuador": "ec", "Perú": "pe", "Chile": "cl", "Paraguay": "py",
                    "Ghana": "gh", "Camerún": "cm", "Túnez": "tn", "Argelia": "dz", "Corea del Sur": "kr", "Australia": "au", "Irán": "ir", "Arabia Saudita": "sa",
                    "Jamaica": "jm", "Panamá": "pa", "Honduras": "hn", "El Salvador": "sv", "Nigeria": "ng", "Costa de Marfil": "ci", "Egipto": "eg", "Mali": "ml",
                    "Suiza": "ch", "Dinamarca": "dk", "Serbia": "rs", "Ucrania": "ua", "Polonia": "pl", "Suecia": "se", "Escocia": "gb-sct", "Gales": "gb-wls"
                }
                for grupo, selecciones in mapa_grupos.items():
                    for sel in selecciones:
                        iso = iso_mapping.get(sel, "un")
                        cursor.execute('INSERT INTO equipos (nombre, zona, archivo_bandera) VALUES (%s, %s, %s) ON CONFLICT (nombre) DO NOTHING', (sel, grupo, f"https://flagcdn.com/w40/{iso}.png"))
                
                cursor.execute("SELECT id_equipo, zona FROM equipos ORDER BY zona, id_equipo")
                eqs = cursor.fetchall()
                zonas = {}
                for eq in eqs: zonas.setdefault(eq[1], []).append(eq[0])
                partidos = []
                for zona, ids in zonas.items():
                    if len(ids) == 4:
                        cruces = [(ids[0], ids[1]), (ids[2], ids[3]), (ids[0], ids[2]), (ids[1], ids[3]), (ids[0], ids[3]), (ids[1], ids[2])]
                        for l, v in cruces: partidos.append((f"Grupo {zona}", l, v))
                
                query_insert_partidos = '''
                    INSERT INTO partidos (fase, id_equipo_local, id_equipo_visitante) 
                    VALUES (%s, %s, %s) 
                    ON CONFLICT (fase, id_equipo_local, id_equipo_visitante) DO NOTHING
                '''
                cursor.executemany(query_insert_partidos, partidos)
                
            conn.commit()
            cursor.close()
        except Exception:
            conn.rollback()
            
        conn.close()

    @staticmethod
    def get_config():
        conn = DatabaseManager.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT pts_prode, pts_exacto, pts_parcial, pts_dif, api_key, id_liga, admin_pass, pts_prode_ko, pts_exacto_ko, pts_parcial_ko, pts_dif_ko, fecha_limite FROM configuracion WHERE id = 1")
        datos = cursor.fetchone()
        conn.close()
        return datos

    @staticmethod
    def set_config(p_prode, p_exact, p_parcial, p_dif, api_key, id_liga, admin_pass, p_prode_ko, p_exact_ko, p_parcial_ko, p_dif_ko, fecha_limite):
        conn = DatabaseManager.get_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE configuracion SET pts_prode=%s, pts_exacto=%s, pts_parcial=%s, pts_dif=%s, api_key=%s, id_liga=%s, admin_pass=%s, pts_prode_ko=%s, pts_exacto_ko=%s, pts_parcial_ko=%s, pts_dif_ko=%s, fecha_limite=%s WHERE id=1", (p_prode, p_exact, p_parcial, p_dif, api_key, id_liga, admin_pass, p_prode_ko, p_exact_ko, p_parcial_ko, p_dif_ko, fecha_limite))
        conn.commit()
        conn.close()

    @staticmethod
    def get_equipos_lista():
        conn = DatabaseManager.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id_equipo, nombre FROM equipos ORDER BY nombre")
        datos = cursor.fetchall()
        conn.close()
        return datos

    # --- REPARADO PARA POSTGRESQL (COMILLAS DOBLES EN ALIAS) ---
    @staticmethod
    def get_partidos_con_nombres():
        conn = DatabaseManager.get_connection()
        cursor = conn.cursor()
        query = '''
            SELECT 
                p.id_partido as "id_partido", p.fase as "Fase", el.archivo_bandera as " ", el.nombre as "Local", 
                p.goles_local as "GL Real", p.goles_visitante as "GV Real", ev.nombre as "Visitante", ev.archivo_bandera as "  "
            FROM partidos p 
            JOIN equipos el ON p.id_equipo_local = el.id_equipo 
            JOIN equipos ev ON p.id_equipo_visitante = ev.id_equipo 
            ORDER BY 
                CASE 
                    WHEN p.fase LIKE 'Grupo%%' THEN 1 WHEN p.fase = '16avos' THEN 2 WHEN p.fase = '8vos' THEN 3
                    WHEN p.fase = '4tos' THEN 4 WHEN p.fase = 'Semi' THEN 5 WHEN p.fase = 'Final' THEN 6 ELSE 7 
                END, p.id_partido
        '''
        cursor.execute(query)
        columns = [desc[0] for desc in cursor.description]
        df = pd.DataFrame(cursor.fetchall(), columns=columns)
        conn.close()
        return df

    @staticmethod
    def guardar_resultado(id_partido, goles_l, goles_v):
        conn = DatabaseManager.get_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE partidos SET goles_local = %s, goles_visitante = %s WHERE id_partido = %s", (goles_l, goles_v, id_partido))
        conn.commit()
        conn.close()

    @staticmethod
    def importar_resultados_excel_admin(df):
        conn = DatabaseManager.get_connection()
        cursor = conn.cursor()
        for _, row in df.iterrows():
            try:
                id_partido = int(row['ID_Partido'])
                if pd.isna(row['Goles_L']) or pd.isna(row['Goles_V']): continue
                cursor.execute("UPDATE partidos SET goles_local = %s, goles_visitante = %s WHERE id_partido = %s", (int(row['Goles_L']), int(row['Goles_V']), id_partido))
            except: continue
        conn.commit()
        conn.close()

    @staticmethod
    def get_usuarios():
        conn = DatabaseManager.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT nombre FROM usuarios ORDER BY nombre")
        datos = [r[0] for r in cursor.fetchall()]
        conn.close()
        return datos

    @staticmethod
    def importar_apuestas_excel(df):
        conn = DatabaseManager.get_connection()
        cursor = conn.cursor()
        for _, row in df.iterrows():
            try: id_partido = int(row['ID_Partido'])
            except: continue
            competidor = str(row['Competidor']).strip()
            cursor.execute("INSERT INTO usuarios (nombre) VALUES (%s) ON CONFLICT (nombre) DO NOTHING", (competidor,))
            cursor.execute("SELECT id_usuario FROM usuarios WHERE nombre = %s", (competidor,))
            id_usr = cursor.fetchone()[0]
            eq_l, eq_v = str(row['Local']).strip(), str(row['Visitante']).strip()
            try: goles_l, goles_v = int(row['Goles_L']), int(row['Goles_V'])
            except: continue
            cursor.execute('''INSERT INTO apuestas (id_usuario, id_partido, apuesta_goles_local, apuesta_goles_visitante, equipo_l_predicho, equipo_v_predicho) VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT(id_usuario, id_partido) DO UPDATE SET apuesta_goles_local=EXCLUDED.apuesta_goles_local, apuesta_goles_visitante=EXCLUDED.apuesta_goles_visitante, equipo_l_predicho=EXCLUDED.equipo_l_predicho, equipo_v_predicho=EXCLUDED.equipo_v_predicho''', (id_usr, id_partido, goles_l, goles_v, eq_l, eq_v))
        conn.commit()
        conn.close()

    @staticmethod
    def calcular_ranking_avanzado():
        conn = DatabaseManager.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT pts_prode, pts_exacto, pts_parcial, pts_dif, pts_prode_ko, pts_exacto_ko, pts_parcial_ko, pts_dif_ko FROM configuracion WHERE id = 1")
        cfg_pts = cursor.fetchone()
        
        cursor.execute('''SELECT p.id_partido, p.fase, el.nombre, ev.nombre, p.goles_local, p.goles_visitante FROM partidos p JOIN equipos el ON p.id_equipo_local = el.id_equipo JOIN equipos ev ON p.id_equipo_visitante = ev.id_equipo WHERE p.goles_local IS NOT NULL''')
        partidos_jugados = cursor.fetchall()
        
        for id_partido, fase, real_l_nombre, real_v_nombre, real_l_gol, real_v_gol in partidos_jugados:
            cursor.execute("SELECT id_usuario, apuesta_goles_local, apuesta_goles_visitante, equipo_l_predicho, equipo_v_predicho FROM apuestas WHERE id_partido = %s AND apuesta_goles_local IS NOT NULL", (id_partido,))
            apuestas = cursor.fetchall()
            res_real = 1 if real_l_gol > real_v_gol else (0 if real_l_gol == real_v_gol else -1)
            
            if "Grupo" in fase: pt_p, pt_ex, pt_pa, pt_df = float(cfg_pts[0]), float(cfg_pts[1]), float(cfg_pts[2]), float(cfg_pts[3])
            else: pt_p, pt_ex, pt_pa, pt_df = float(cfg_pts[4]), float(cfg_pts[5]), float(cfg_pts[6]), float(cfg_pts[7])
                
            for id_user, ap_l_gol, ap_v_gol, ap_l_nombre, ap_v_nombre in apuestas:
                res_ap = 1 if ap_l_gol > ap_v_gol else (0 if ap_l_gol == ap_v_gol else -1)
                puntos_base = 0.0
                if res_real == res_ap:
                    puntos_base += pt_p
                    if real_l_gol == ap_l_gol and real_v_gol == ap_v_gol: puntos_base += pt_ex
                    else:
                        if real_l_gol == ap_l_gol or real_v_gol == ap_v_gol: puntos_base += pt_pa
                        if (real_l_gol - real_v_gol) == (ap_l_gol - ap_v_gol): puntos_base += pt_df
                cursor.execute("UPDATE apuestas SET puntos_obtenidos = %s WHERE id_usuario = %s AND id_partido = %s", (puntos_base, id_user, id_partido))
        conn.commit()
        
        cursor.execute('SELECT u.nombre as "Competidor", COALESCE(SUM(a.puntos_obtenidos), 0) as "Puntos" FROM usuarios u LEFT JOIN apuestas a ON u.id_usuario = a.id_usuario GROUP BY u.id_usuario, u.nombre ORDER BY "Puntos" DESC, u.nombre ASC')
        columns = [desc[0] for desc in cursor.description]
        df_rank = pd.DataFrame(cursor.fetchall(), columns=columns)
        conn.close()
        return df_rank

    # --- REPARADO PARA POSTGRESQL (COMILLAS DOBLES EN ALIAS) ---
    @staticmethod
    def get_apuestas_usuario_web(nombre_usuario):
        conn = DatabaseManager.get_connection()
        cursor = conn.cursor()
        query = '''SELECT p.fase as "Fase", el.archivo_bandera as " ", a.equipo_l_predicho as "Local", a.apuesta_goles_local as "GL Pred", a.apuesta_goles_visitante as "GV Pred", a.equipo_v_predicho as "Visitante", ev.archivo_bandera as "  ", p.goles_local as "GL Real", p.goles_visitante as "GV Real", a.puntos_obtenidos as "Pts Ganados" FROM apuestas a JOIN usuarios u ON a.id_usuario = u.id_usuario JOIN partidos p ON a.id_partido = p.id_partido JOIN equipos el ON p.id_equipo_local = el.id_equipo JOIN equipos ev ON p.id_equipo_visitante = ev.id_equipo WHERE u.nombre = %s'''
        cursor.execute(query, (nombre_usuario,))
        columns = [desc[0] for desc in cursor.description]
        df = pd.DataFrame(cursor.fetchall(), columns=columns)
        conn.close()
        return df

    @staticmethod
    def obtener_datos_auditoria_puntos():
        conn = DatabaseManager.get_connection()
        cursor = conn.cursor()
        query = '''SELECT u.nombre as "Competidor", p.fase as "Fase", el.nombre || ' vs ' || ev.nombre as "Partido", a.apuesta_goles_local as "Al", a.apuesta_goles_visitante as "Av", p.goles_local as "Rl", p.goles_visitante as "Rv" FROM apuestas a JOIN usuarios u ON a.id_usuario = u.id_usuario JOIN partidos p ON a.id_partido = p.id_partido JOIN equipos el ON p.id_equipo_local = el.id_equipo JOIN equipos ev ON p.id_equipo_visitante = ev.id_equipo WHERE p.goles_local IS NOT NULL AND a.apuesta_goles_local IS NOT NULL'''
        cursor.execute(query)
        columns = [desc[0] for desc in cursor.description]
        df = pd.DataFrame(cursor.fetchall(), columns=columns)
        conn.close()
        return df

# --- INICIALIZACIÓN ---
DatabaseManager.init_db()

tabs = st.tabs(["📊 Posiciones y Apuestas", "📤 Subir Mis Apuestas", "⚙️ Panel Administrador"])

# ==========================================
# TAB 1: DASHBOARD PÚBLICO
# ==========================================
with tabs[0]:
    cfg = DatabaseManager.get_config()
    with st.expander("📜 Ver Reglamento y Sistema de Puntuación"):
        st.subheader("📝 Cálculo de Puntos Automático")
        st.write("Los ítems suman de forma independiente según la fase del torneo:")
        col_r1, col_r2 = st.columns(2)
        with col_r1: st.markdown(f"**🏟️ Fase de Grupos:**\n* Ganador/Empate: **+{cfg[0]} Pts**\n* Exacto: **+{cfg[1]} Pts**\n* Goles de un equipo: **+{cfg[2]} Pts**\n* Diferencia de goles: **+{cfg[3]} Pts**")
        with col_r2: st.markdown(f"**⚔️ Segunda Vuelta (KO):**\n* Ganador KO: **+{cfg[7]} Pts**\n* Exacto KO: **+{cfg[8]} Pts**\n* Goles KO: **+{cfg[9]} Pts**\n* Diferencia KO: **+{cfg[10]} Pts**")
            
    st.markdown("---")
    col1, col2 = st.columns([1, 1.2])
    with col1:
        st.subheader("⭐ Tabla de Posiciones")
        df_rank = DatabaseManager.calcular_ranking_avanzado()
        if not df_rank.empty:
            df_rank.index = df_rank.index + 1
            st.dataframe(df_rank, use_container_width=True)
        else: st.info("No hay puntuaciones registradas.")
    with col2:
        st.subheader("📅 Fixture y Resultados Oficiales")
        df_public_partidos = DatabaseManager.get_partidos_con_nombres()
        if not df_public_partidos.empty:
            fases_unicas = df_public_partidos['Fase'].unique()
            fase_colors = {fase: 'rgba(30, 144, 255, 0.12)' if idx % 2 == 0 else 'rgba(0, 0, 0, 0)' for idx, fase in enumerate(fases_unicas)}
            
            # --- CORREGIDO: MAPEADO PRECISO DE COLUMNAS CON SUS NUEVAS ETIQUETAS DE POSTGRES ---
            st.dataframe(df_public_partidos.style.apply(lambda r: [f"background-color: {fase_colors.get(r['Fase'], '')}" for _ in r], axis=1), use_container_width=True, hide_index=True, column_config={
                " ": st.column_config.ImageColumn(label=""), 
                "  ": st.column_config.ImageColumn(label=""), 
                "GL Real": st.column_config.NumberColumn(label="GL"), 
                "GV Real": st.column_config.NumberColumn(label="GV")
            })
        else: st.info("El fixture todavía no fue generado.")

    st.markdown("---")
    st.subheader("🔍 Apuestas")
    if datetime.now().strftime("%Y-%m-%d %H:%M:%S") >= cfg[11]:
        usuarios = DatabaseManager.get_usuarios()
        if usuarios:
            user_sel = st.selectbox("Selecciona un competidor:", usuarios)
            if user_sel:
                df_user_ap = DatabaseManager.get_apuestas_usuario_web(user_sel)
                if not df_user_ap.empty:
                    colors_usr = {fase: 'rgba(30, 144, 255, 0.12)' if idx % 2 == 0 else 'rgba(0, 0, 0, 0)' for idx, fase in enumerate(df_user_ap['Fase'].unique())}
                    st.dataframe(df_user_ap.style.apply(lambda r: [f"background-color: {colors_usr.get(r['Fase'], '')}" for _ in r], axis=1), use_container_width=True, hide_index=True, column_config={
                        " ": st.column_config.ImageColumn(label=""), 
                        "  ": st.column_config.ImageColumn(label="")
                    })
        else: st.info("Aún no hay usuarios cargados en el sistema.")
    else: st.warning(f"🔒 **Pronósticos Protegidos:** Las apuestas de todos se revelarán el **{cfg[11]}** para evitar copias.")

# ==========================================
# TAB 2: USER CARGA APUESTAS
# ==========================================
with tabs[1]:
    st.subheader("📝 Envía tus Pronósticos")
    st.write("Descarga la plantilla oficial, escribe tu nombre en la celda amarilla B3, completa tus goles y subila acá.")
    df_partidos = DatabaseManager.get_partidos_con_nombres()
    if not df_partidos.empty:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Apuestas"; ws.views.sheetView[0].showGridLines = True
        ws.merge_cells("A1:F1"); ws["A1"] = "PRODE MUNDIAL - FORMULARIO"
        ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF"); ws["A1"].fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 35
        ws["A3"] = "COMPETIDOR (Nombre):"; ws["A3"].font = Font(name="Arial", size=11, bold=True)
        ws["B3"] = "[Escriba su nombre aquí]"; ws["B3"].fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
        thin = Border(left=Side(style='thin', color='A9A9A9'), right=Side(style='thin', color='A9A9A9'), top=Side(style='thin', color='A9A9A9'), bottom=Side(style='thin', color='A9A9A9'))
        thick_bot = Border(left=Side(style='thin', color='A9A9A9'), right=Side(style='thin', color='A9A9A9'), top=Side(style='thin', color='A9A9A9'), bottom=Side(style='medium', color='000000'))
        ws["B3"].border = thin
        
        for col_num, h in enumerate(['ID_Partido', 'Fase', 'Local', 'Goles_L', 'Goles_V', 'Visitante'], 1):
            c = ws.cell(row=5, column=col_num, value=h); c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF"); c.fill = PatternFill(start_color="4682B4", end_color="4682B4", fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin
        ws.row_dimensions[5].height = 30
        
        row_num = 6; prev_fase = None
        for i, r in df_partidos.iterrows():
            if prev_fase and r['Fase'] != prev_fase:
                for col_num in range(1, 7): ws.cell(row=row_num-1, column=col_num).border = thick_bot
            ws.row_dimensions[row_num].height = 24
            ws.cell(row=row_num, column=1, value=r['id_partido']).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=2, value=r['Fase']).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=3, value=r['Local'])
            ws.cell(row=row_num, column=6, value=r['Visitante'])
            for col_num in range(1, 7):
                cell = ws.cell(row=row_num, column=col_num); cell.font = Font(name="Arial", size=11)
                if cell.border != thick_bot: cell.border = thin
                if i % 2 == 1 and col_num not in [4, 5]: cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            prev_fase = r['Fase']; row_num += 1
            
        for col_num in range(1, 7): ws.cell(row=row_num-1, column=col_num).border = thick_bot
        ws.column_dimensions['A'].width = 12; ws.column_dimensions['B'].width = 18; ws.column_dimensions['C'].width = 25; ws.column_dimensions['D'].width = 14; ws.column_dimensions['E'].width = 14; ws.column_dimensions['F'].width = 25
        out = io.BytesIO(); wb.save(out)
        
        st.download_button(label="📥 Descargar Plantilla de Apuestas", data=out.getvalue(), file_name="Plantilla_Apuestas_Mundial.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.markdown("---")
        uploaded_file = st.file_uploader("Sube tu Excel completado aquí:", type=["xlsx"])
        if uploaded_file is not None:
            try:
                df_meta = pd.read_excel(uploaded_file, nrows=3, header=None)
                competidor = str(df_meta.iloc[2, 1]).strip()
                if not competidor or competidor == "nan" or competidor == "[Escriba su nombre aquí]": st.error("Error: Olvidaste escribir tu nombre en la celda amarilla B3.")
                else:
                    DatabaseManager.importar_apuestas_excel(pd.read_excel(uploaded_file, skiprows=4).assign(Competidor=competidor))
                    st.success(f"¡Excelente! Pronósticos de '{competidor}' cargados correctamente.")
            except Exception as e: st.error(f"Error procesando el archivo: {e}")
    else: st.warning("El fixture no fue generado aún.")

# ==========================================
# TAB 3: ADMIN PANEL (PROTEGIDO)
# ==========================================
with tabs[2]:
    cfg = DatabaseManager.get_config()
    pass_input = st.text_input("Ingresa la Contraseña de Administrador:", type="password")
    if pass_input == cfg[6]:
        st.success("Conexión Segura con PostgreSQL Cloud En Línea")
        thin_b = Border(left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'), top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3'))
        
        st.markdown("### 📊 Informes de Auditoría de Producción")
        df_auditoria = DatabaseManager.obtener_datos_auditoria_puntos()
        if not df_auditoria.empty:
            wb_aud = openpyxl.Workbook(); ws_aud = wb_aud.active; ws_aud.title = "Auditoría"; ws_aud.views.sheetView[0].showGridLines = True
            ws_aud.merge_cells("A1:J1"); ws_aud["A1"] = "INFORME DETALLADO DE DESGLOSE DE PUNTOS"
            ws_aud["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF"); ws_aud["A1"].fill = PatternFill(start_color="1F4E5B", end_color="1F4E5B", fill_type="solid")
            ws_aud["A1"].alignment = Alignment(horizontal="center", vertical="center"); ws_aud.row_dimensions[1].height = 35
            
            for col_idx, h in enumerate(['Competidor', 'Fase', 'Partido', 'Pronóstico', 'Resultado Real', 'Pts Ganador', 'Pts Exacto', 'Pts Goles', 'Pts Diferencia', 'Total Partido'], 1):
                cell = ws_aud.cell(row=3, column=col_idx, value=h); cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF"); cell.fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid"); cell.alignment = Alignment(horizontal="center"); cell.border = thin_b
            ws_aud.row_dimensions[3].height = 28
            
            r_idx = 4
            for _, row in df_auditoria.iterrows():
                try: al, av, rl, rv = int(row['Al']), int(row['Av']), int(row['Rl']), int(row['Rv'])
                except: continue
                res_real = 1 if rl > rv else (0 if rl == rv else -1)
                res_ap = 1 if al > av else (0 if al == av else -1)
                pt_p, pt_ex, pt_pa, pt_df = (float(cfg[0]), float(cfg[1]), float(cfg[2]), float(cfg[3])) if "Grupo" in row['Fase'] else (float(cfg[7]), float(cfg[8]), float(cfg[9]), float(cfg[10]))
                
                p_win, p_ex, p_gol, p_df = 0.0, 0.0, 0.0, 0.0
                if res_real == res_ap:
                    p_win = pt_p
                    if rl == al and rv == av: p_ex = pt_ex
                    else:
                        if rl == al or rv == av: p_gol = pt_pa
                        if (rl - rv) == (al - av): p_df = pt_df
                
                ws_aud.cell(row=r_idx, column=1, value=row['Competidor'])
                ws_aud.cell(row=r_idx, column=2, value=row['Fase']).alignment = Alignment(horizontal="center")
                ws_aud.cell(row=r_idx, column=3, value=row['Partido'])
                ws_aud.cell(row=r_idx, column=4, value=f"{al} - {av}").alignment = Alignment(horizontal="center")
                ws_aud.cell(row=r_idx, column=5, value=f"{rl} - {rv}").alignment = Alignment(horizontal="center")
                ws_aud.cell(row=r_idx, column=6, value=p_win).alignment = Alignment(horizontal="center")
                ws_aud.cell(row=r_idx, column=7, value=p_ex).alignment = Alignment(horizontal="center")
                ws_aud.cell(row=r_idx, column=8, value=p_gol).alignment = Alignment(horizontal="center")
                ws_aud.cell(row=r_idx, column=9, value=p_df).alignment = Alignment(horizontal="center")
                c_tot = ws_aud.cell(row=r_idx, column=10, value=p_win + p_ex + p_gol + p_df); c_tot.font = Font(bold=True); c_tot.alignment = Alignment(horizontal="center"); c_tot.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
                for c in range(1, 11): ws_aud.cell(row=r_idx, column=c).border = thin_b
                r_idx += 1
            for col in ws_aud.columns: ws_aud.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max(len(str(cell.value or '')) for cell in col) + 3, 12)
            out_aud = io.BytesIO(); wb_aud.save(out_aud)
            st.download_button(label="📊 Descargar Excel: Desglose de Puntos", data=out_aud.getvalue(), file_name="Desglose_Puntos_Usuarios.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else: st.caption("El desglose de puntos estará disponible cuando empiece el torneo.")
        
        st.markdown("---")
        st.markdown("### 🔄 Carga de Resultados Oficiales del Torneo")
        col_res1, col_res2 = st.columns(2)
        with col_res1:
            st.markdown("##### 🌐 Opción Principal: Sincronización Automática")
            if st.button("🔄 Sincronizar Resultados vía API Now", use_container_width=True):
                if not cfg[4]: st.error("Error: Falta tu API Key.")
                else:
                    try:
                        res = requests.get("https://v3.football.api-sports.io/fixtures", headers={"x-rapidapi-key": cfg[4], "x-rapidapi-host": "v3.football.api-sports.io"}, params={"league": cfg[5], "season": "2026"})
                        count = 0; df_p = DatabaseManager.get_partidos_con_nombres()
                        for f in res.json().get("response", []):
                            if f["fixture"]["status"]["short"] in ["FT", "AET", "PEN"]:
                                al, av, gl, gv = f["teams"]["home"]["name"].lower(), f["teams"]["away"]["name"].lower(), f["goals"]["home"], f["goals"]["away"]
                                for idx, row in df_p.iterrows():
                                    if al in row['Local'].lower() and av in row['Visitante'].lower(): DatabaseManager.guardar_resultado(int(row['id_partido']), gl, gv); count += 1
                        st.success(f"¡Sincronización terminada! {count} partidos actualizados."); st.rerun()
                    except Exception as e: st.error(f"Error de API: {e}")
        with col_res2:
            st.markdown("##### 📂 Opción Secundaria: Contingencia Manual por Excel")
            df_actual_goles = DatabaseManager.get_partidos_con_nombres()
            if not df_actual_goles.empty:
                wb_adm_res = openpyxl.Workbook(); ws_adm_res = wb_adm_res.active; ws_adm_res.title = "Resultados"; ws_adm_res.views.sheetView[0].showGridLines = True
                ws_adm_res.merge_cells("A1:F1"); ws_adm_res["A1"] = "PRODE MUNDIAL - CARGA MANUAL DE RESULTADOS"
                ws_adm_res["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFF"); ws_adm_res["A1"].fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
                ws_adm_res["A1"].alignment = Alignment(horizontal="center", vertical="center"); ws_adm_res.row_dimensions[1].height = 35
                for col_num, h in enumerate(['ID_Partido', 'Fase', 'Local', 'Goles_L', 'Goles_V', 'Visitante'], 1):
                    c = ws_adm_res.cell(row=5, column=col_num, value=h); c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF"); c.fill = PatternFill(start_color="343a40", end_color="343a40", fill_type="solid"); c.alignment = Alignment(horizontal="center"); c.border = thin_b
                
                row_n = 6
                for _, r in df_actual_goles.iterrows():
                    ws_adm_res.cell(row=row_n, column=1, value=r['id_partido']).alignment = Alignment(horizontal="center")
                    ws_adm_res.cell(row=row_n, column=2, value=r['Fase']).alignment = Alignment(horizontal="center")
                    ws_adm_res.cell(row=row_n, column=3, value=r['Local'])
                    ws_adm_res.cell(row=row_n, column=4, value="" if pd.isna(r['GL Real']) else int(r['GL Real'])).alignment = Alignment(horizontal="center")
                    ws_adm_res.cell(row=row_n, column=5, value="" if pd.isna(r['GV Real']) else int(r['GV Real'])).alignment = Alignment(horizontal="center")
                    ws_adm_res.cell(row=row_n, column=6, value=r['Visitante'])
                    for c in range(1, 7):
                        ws_adm_res.cell(row=row_n, column=c).border = thin_b
                        if c in [4, 5]: ws_adm_res.cell(row=row_n, column=c).fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
                    row_n += 1
                ws_adm_res.column_dimensions['A'].width = 12; ws_adm_res.column_dimensions['B'].width = 15; ws_adm_res.column_dimensions['C'].width = 22; ws_adm_res.column_dimensions['D'].width = 12; ws_adm_res.column_dimensions['E'].width = 12; ws_adm_res.column_dimensions['F'].width = 22
                out_adm_res = io.BytesIO(); wb_adm_res.save(out_adm_res)
                st.download_button(label="📥 Descargar Planilla de Resultados Reales", data=out_adm_res.getvalue(), file_name="Planilla_Resultados_Oficiales.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                uploaded_res_file = st.file_uploader("Subir Excel de Resultados Oficiales:", type=["xlsx"], key="admin_res_upload")
                if uploaded_res_file is not None:
                    try:
                        DatabaseManager.importar_resultados_excel_admin(pd.read_excel(uploaded_res_file, skiprows=4))
                        st.success("¡Resultados cargados desde Excel exitosamente!"); st.rerun()
                    except Exception as e: st.error(f"Error procesando el Excel: {e}")
        
        st.markdown("---")
        col_adm1, col_adm2 = st.columns(2)
        with col_adm1:
            st.subheader("### Habilitar Eliminación Directa (Mano a Mano)")
            lista_eq = DatabaseManager.get_equipos_lista()
            if lista_eq:
                fase_sel = st.selectbox("Fase:", ["16avos", "8vos", "4tos", "Semi", "Final"])
                eq_loc_id = st.selectbox("Selecciona Equipo Local:", [e[0] for e in lista_eq], format_func=lambda x: next(i[1] for i in lista_eq if i[0] == x))
                eq_vis_id = st.selectbox("Selecciona Equipo Visitante:", [e[0] for e in lista_eq], format_func=lambda x: next(i[1] for i in lista_eq if i[0] == x), key="vis")
                if st.button("➕ Publicar Cruce de Eliminación", use_container_width=True):
                    if eq_loc_id == eq_vis_id: st.error("Un equipo no puede jugar contra sí mismo.")
                    else: DatabaseManager.insertar_partido_manual(fase_sel, eq_loc_id, eq_vis_id); st.success(f"Partido de {fase_sel} publicado."); st.rerun()
                        
        with col_adm2:
            st.subheader("⚙️ Configuración de Puntuación y Sistema")
            st.markdown("##### 🏟️ Fase de Grupos / ⚔️ Segunda Vuelta KO")
            p_prode = st.number_input("Pts Ganador", value=cfg[0]); p_exact = st.number_input("Pts Exacto", value=cfg[1]); p_parc = st.number_input("Pts Goles", value=cfg[2]); p_dif = st.number_input("Pts Diferencia", value=cfg[3])
            p_prode_ko = st.number_input("Pts Ganador KO", value=cfg[7]); p_exact_ko = st.number_input("Pts Exacto KO", value=cfg[8]); p_parc_ko = st.number_input("Pts Goles KO", value=cfg[9]); p_dif_ko = st.number_input("Pts Diferencia KO", value=cfg[10])
            st.markdown("##### 🔑 Credenciales, Llaves y Fechas")
            ak = st.text_input("API Key de API-Football:", value=cfg[4]); id_l = st.text_input("ID de la Liga/Mundial:", value=cfg[5]); f_limite = st.text_input("Fecha límite (AAAA-MM-DD HH:MM:SS):", value=cfg[11]); new_pass = st.text_input("Nueva Clave Admin:", value=cfg[6])
            if st.button("Guardar Cambios de Configuración", use_container_width=True):
                DatabaseManager.set_config(p_prode, p_exact, p_parc, p_dif, ak, id_l, new_pass, p_prode_ko, p_exact_ko, p_parc_ko, p_dif_ko, f_limite); st.success("Configuración guardada."); st.rerun()
                
        st.markdown("---")
        st.subheader("### Grilla de Partidos Actuales")
        df_adm_partidos = DatabaseManager.get_partidos_con_nombres()
        fases_adm = df_adm_partidos['Fase'].unique()
        fase_colors_adm = {fase: 'rgba(30, 144, 255, 0.12)' if idx % 2 == 0 else 'rgba(0, 0, 0, 0)' for idx, fase in enumerate(fases_adm)}
        st.dataframe(df_adm_partidos.style.apply(lambda r: [f"background-color: {fase_colors_adm.get(r['Fase'], '')}" for _ in r], axis=1), use_container_width=True, hide_index=True, column_config={" ": st.column_config.ImageColumn(label=""), "  ": st.column_config.ImageColumn(label="")})
    else:
        if pass_input: st.error("Contraseña incorrecta.")
